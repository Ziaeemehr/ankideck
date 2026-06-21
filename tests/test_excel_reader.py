"""Tests for read_cards_excel in ankideck.reader."""

import pytest
import openpyxl

from ankideck.reader import read_cards_excel, Card


def _make_xlsx(tmp_path, rows, sheet_name="Sheet"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return str(path)


def test_basic_two_column(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English"],
        ["bonjour", "hello"],
        ["merci", "thank you"],
    ])
    cards = read_cards_excel(path, front_col="French", back_cols=["English"])
    assert len(cards) == 2
    assert cards[0].front == "bonjour"
    assert "hello" in cards[0].back


def test_four_column_structure(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English", "Persian", "Example"],
        ["bonjour", "hello", "سلام", "Bonjour tout le monde."],
    ])
    cards = read_cards_excel(
        path,
        front_col="French",
        back_cols=["English", "Persian", "Example"],
        tts_back_col="Example",
    )
    assert len(cards) == 1
    c = cards[0]
    assert c.front == "bonjour"
    assert "hello" in c.back
    assert "سلام" in c.back
    assert "Bonjour tout le monde" in c.back


def test_tts_front_set(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English"],
        ["chat", "cat"],
    ])
    cards = read_cards_excel(path, front_col="French")
    assert cards[0].tts_front == "chat"


def test_tts_back_set(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English", "Example"],
        ["chat", "cat", "Le chat dort."],
    ])
    cards = read_cards_excel(path, front_col="French", tts_back_col="Example")
    assert cards[0].tts_back == "Le chat dort."


def test_tts_back_empty_when_not_set(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English", "Example"],
        ["chat", "cat", "Le chat dort."],
    ])
    cards = read_cards_excel(path, front_col="French", tts_back_col=None)
    assert cards[0].tts_back == ""


def test_col_labels_in_back_html(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English"],
        ["bonjour", "hello"],
    ])
    cards = read_cards_excel(
        path,
        front_col="French",
        back_cols=["English"],
        col_labels={"English": "EN"},
    )
    assert "<b>EN:</b>" in cards[0].back


def test_default_back_cols_excludes_front(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English", "Persian"],
        ["bonjour", "hello", "سلام"],
    ])
    cards = read_cards_excel(path, front_col="French")
    # back should contain English and Persian but not French again
    assert "hello" in cards[0].back
    assert "سلام" in cards[0].back
    assert cards[0].front not in cards[0].back


def test_example_col_styled_differently(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English", "Example"],
        ["chat", "cat", "Le chat dort."],
    ])
    cards = read_cards_excel(
        path, front_col="French",
        back_cols=["English", "Example"], tts_back_col="Example",
    )
    # Example should be in italic/styled div, not in the label format
    assert "font-style:italic" in cards[0].back
    assert "<b>Example:</b>" not in cards[0].back


def test_skips_empty_front_rows(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English"],
        ["bonjour", "hello"],
        [None, "orphan"],
        ["merci", "thank you"],
    ])
    cards = read_cards_excel(path, front_col="French")
    assert len(cards) == 2


def test_missing_column_raises(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English"],
        ["bonjour", "hello"],
    ])
    with pytest.raises(ValueError, match="Column.*not found"):
        read_cards_excel(path, front_col="German")


def test_tags_attached(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["French", "English"],
        ["bonjour", "hello"],
    ])
    cards = read_cards_excel(path, front_col="French", tags=["a2", "greetings"])
    assert "a2" in cards[0].tags
    assert "greetings" in cards[0].tags


def test_sheet_name(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "ignore"
    ws = wb.create_sheet("Vocab")
    ws.append(["French", "English"])
    ws.append(["chat", "cat"])
    path = tmp_path / "multi.xlsx"
    wb.save(str(path))
    cards = read_cards_excel(str(path), front_col="French", sheet_name="Vocab")
    assert len(cards) == 1
